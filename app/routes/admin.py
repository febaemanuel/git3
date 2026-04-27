"""Admin dashboard and exports."""

import logging
from datetime import datetime, timedelta

from flask import (
    Blueprint, flash, jsonify, redirect, render_template, request,
    send_file, url_for,
)
from flask_login import current_user, login_required
from io import BytesIO
import pandas as pd

from app.extensions import db
from app.models import (
    AgendamentoConsulta, Campanha, CampanhaConsulta, ConfigWhatsApp,
    Contato, LogMsg, LogMsgConsulta, PesquisaSatisfacao,
    RespostaAutomatica, Telefone, TicketAtendimento, Usuario,
)
from app.services.whatsapp import WhatsApp
from app.utils import admin_required, get_dashboard_route


logger = logging.getLogger(__name__)


bp = Blueprint('admin', __name__)


@bp.route('/admin/dashboard')
@login_required
@admin_required
def dashboard():
    """Dashboard administrativo com métricas globais do sistema"""
    from sqlalchemy import func, case

    # Parâmetro de modo selecionado
    modo_selecionado = request.args.get('modo', 'todos')  # 'consulta', 'fila' ou 'todos'

    # =====================================================================
    # ESTATÍSTICAS DE USUÁRIOS (filtradas por modo)
    # =====================================================================
    if modo_selecionado == 'consulta':
        total_usuarios = Usuario.query.filter_by(tipo_sistema='AGENDAMENTO_CONSULTA').count()
        usuarios_ativos = Usuario.query.filter_by(ativo=True, tipo_sistema='AGENDAMENTO_CONSULTA').count()
    elif modo_selecionado == 'fila':
        total_usuarios = Usuario.query.filter_by(tipo_sistema='BUSCA_ATIVA').count()
        usuarios_ativos = Usuario.query.filter_by(ativo=True, tipo_sistema='BUSCA_ATIVA').count()
    else:
        total_usuarios = Usuario.query.count()
        usuarios_ativos = Usuario.query.filter_by(ativo=True).count()

    usuarios_admin = Usuario.query.filter_by(is_admin=True).count()

    # Usuários por tipo de sistema
    usuarios_fila = Usuario.query.filter_by(tipo_sistema='BUSCA_ATIVA').count()
    usuarios_consulta = Usuario.query.filter_by(tipo_sistema='AGENDAMENTO_CONSULTA').count()

    # =====================================================================
    # ESTATÍSTICAS DO MODO FILA (BUSCA_ATIVA)
    # =====================================================================
    total_campanhas_fila = Campanha.query.count()
    campanhas_fila_ativas = Campanha.query.filter(Campanha.status.in_(['validando', 'pronta', 'enviando'])).count()

    # Totais agregados de todas as campanhas de fila
    stats_fila = db.session.query(
        func.coalesce(func.sum(Campanha.total_contatos), 0).label('total_contatos'),
        func.coalesce(func.sum(Campanha.total_enviados), 0).label('total_enviados'),
        func.coalesce(func.sum(Campanha.total_confirmados), 0).label('total_confirmados'),
        func.coalesce(func.sum(Campanha.total_rejeitados), 0).label('total_rejeitados'),
        func.coalesce(func.sum(Campanha.total_erros), 0).label('total_erros')
    ).first()

    fila_total_contatos = stats_fila.total_contatos or 0
    fila_total_enviados = stats_fila.total_enviados or 0
    fila_total_confirmados = stats_fila.total_confirmados or 0
    fila_total_rejeitados = stats_fila.total_rejeitados or 0
    fila_total_erros = stats_fila.total_erros or 0

    # Taxa de confirmação da fila
    fila_taxa_confirmacao = round((fila_total_confirmados / fila_total_enviados * 100), 1) if fila_total_enviados > 0 else 0

    # =====================================================================
    # ESTATÍSTICAS DO MODO CONSULTA (AGENDAMENTO_CONSULTA)
    # =====================================================================
    total_campanhas_consulta = CampanhaConsulta.query.count()
    campanhas_consulta_ativas = CampanhaConsulta.query.filter(CampanhaConsulta.status.in_(['pronta', 'enviando'])).count()

    # Totais agregados de todas as campanhas de consulta
    stats_consulta = db.session.query(
        func.coalesce(func.sum(CampanhaConsulta.total_consultas), 0).label('total_consultas'),
        func.coalesce(func.sum(CampanhaConsulta.total_enviados), 0).label('total_enviados'),
        func.coalesce(func.sum(CampanhaConsulta.total_confirmados), 0).label('total_confirmados'),
        func.coalesce(func.sum(CampanhaConsulta.total_rejeitados), 0).label('total_rejeitados')
    ).first()

    consulta_total = stats_consulta.total_consultas or 0
    consulta_enviados = stats_consulta.total_enviados or 0
    consulta_confirmados = stats_consulta.total_confirmados or 0
    consulta_rejeitados = stats_consulta.total_rejeitados or 0

    # Taxa de confirmação de consultas
    consulta_taxa_confirmacao = round((consulta_confirmados / consulta_enviados * 100), 1) if consulta_enviados > 0 else 0

    # =====================================================================
    # ESTATÍSTICAS DETALHADAS - AGENDAMENTO DE CONSULTAS
    # =====================================================================
    # Disparos de MSG 1 (agendamentos enviados)
    total_disparos_msg1 = AgendamentoConsulta.query.filter_by(mensagem_enviada=True).count()

    # Comprovantes enviados (status CONFIRMADO = comprovante foi enviado)
    total_comprovantes_enviados = AgendamentoConsulta.query.filter_by(status='CONFIRMADO').count()

    # Aguardando comprovante
    total_aguardando_comprovante = AgendamentoConsulta.query.filter_by(status='AGUARDANDO_COMPROVANTE').count()

    # Cancelados (sem resposta)
    total_cancelados = AgendamentoConsulta.query.filter_by(status='CANCELADO').count()

    # =====================================================================
    # DESEMPENHO POR USUÁRIO (MODO CONSULTA)
    # =====================================================================
    # Filtrar por modo se necessário
    desempenho_query = db.session.query(
        Usuario.id,
        Usuario.nome,
        Usuario.tipo_sistema,
        # MSG 1 enviadas (disparos)
        func.sum(case((AgendamentoConsulta.mensagem_enviada == True, 1), else_=0)).label('disparos_msg1'),
        # Comprovantes enviados
        func.sum(case((AgendamentoConsulta.status == 'CONFIRMADO', 1), else_=0)).label('comprovantes_enviados'),
        # Aguardando comprovante
        func.sum(case((AgendamentoConsulta.status == 'AGUARDANDO_COMPROVANTE', 1), else_=0)).label('aguardando_comprovante'),
        # Rejeitados
        func.sum(case((AgendamentoConsulta.status == 'REJEITADO', 1), else_=0)).label('rejeitados'),
        # Cancelados
        func.sum(case((AgendamentoConsulta.status == 'CANCELADO', 1), else_=0)).label('cancelados'),
        # Total de agendamentos
        func.count(AgendamentoConsulta.id).label('total_agendamentos')
    ).join(CampanhaConsulta, CampanhaConsulta.criador_id == Usuario.id
    ).join(AgendamentoConsulta, AgendamentoConsulta.campanha_id == CampanhaConsulta.id
    )

    # Aplicar filtro de modo
    if modo_selecionado == 'consulta':
        desempenho_query = desempenho_query.filter(Usuario.tipo_sistema == 'AGENDAMENTO_CONSULTA')
    elif modo_selecionado == 'fila':
        desempenho_query = desempenho_query.filter(Usuario.tipo_sistema == 'BUSCA_ATIVA')

    desempenho_usuarios = desempenho_query.group_by(Usuario.id, Usuario.nome, Usuario.tipo_sistema
    ).having(func.count(AgendamentoConsulta.id) > 0
    ).order_by(func.sum(case((AgendamentoConsulta.mensagem_enviada == True, 1), else_=0)).desc()
    ).all()

    # =====================================================================
    # MENSAGENS POR USUÁRIO
    # =====================================================================
    # Modo Consulta - mensagens enviadas por usuário (via criador da campanha)
    msgs_por_usuario_consulta = db.session.query(
        Usuario.nome,
        Usuario.id.label('usuario_id'),
        func.count(LogMsgConsulta.id).label('total_msgs'),
        func.sum(case((LogMsgConsulta.direcao == 'enviada', 1), else_=0)).label('enviadas'),
        func.sum(case((LogMsgConsulta.direcao == 'recebida', 1), else_=0)).label('recebidas')
    ).join(CampanhaConsulta, CampanhaConsulta.id == LogMsgConsulta.campanha_id
    ).join(Usuario, Usuario.id == CampanhaConsulta.criador_id
    ).group_by(Usuario.id, Usuario.nome
    ).order_by(func.count(LogMsgConsulta.id).desc()
    ).limit(15).all()

    # Modo Fila - mensagens enviadas por usuário
    msgs_por_usuario_fila = db.session.query(
        Usuario.nome,
        Usuario.id.label('usuario_id'),
        func.count(LogMsg.id).label('total_msgs'),
        func.sum(case((LogMsg.direcao == 'enviada', 1), else_=0)).label('enviadas'),
        func.sum(case((LogMsg.direcao == 'recebida', 1), else_=0)).label('recebidas')
    ).join(Campanha, Campanha.id == LogMsg.campanha_id
    ).join(Usuario, Usuario.id == Campanha.criador_id
    ).group_by(Usuario.id, Usuario.nome
    ).order_by(func.count(LogMsg.id).desc()
    ).limit(15).all()

    # =====================================================================
    # TAXA DE CONFIRMAÇÃO POR USUÁRIO (MODO CONSULTA)
    # =====================================================================
    confirmacao_por_usuario = db.session.query(
        Usuario.nome,
        func.count(AgendamentoConsulta.id).label('total'),
        func.sum(case((AgendamentoConsulta.status == 'CONFIRMADO', 1), else_=0)).label('confirmados'),
        func.sum(case((AgendamentoConsulta.status == 'REJEITADO', 1), else_=0)).label('rejeitados'),
        func.sum(case((AgendamentoConsulta.mensagem_enviada == True, 1), else_=0)).label('enviados')
    ).join(CampanhaConsulta, CampanhaConsulta.id == AgendamentoConsulta.campanha_id
    ).join(Usuario, Usuario.id == CampanhaConsulta.criador_id
    ).group_by(Usuario.id, Usuario.nome
    ).having(func.sum(case((AgendamentoConsulta.mensagem_enviada == True, 1), else_=0)) > 0
    ).order_by(func.count(AgendamentoConsulta.id).desc()
    ).limit(15).all()

    # =====================================================================
    # PESQUISAS DE SATISFAÇÃO
    # =====================================================================
    # Total de pesquisas
    total_pesquisas = PesquisaSatisfacao.query.count()
    pesquisas_respondidas = PesquisaSatisfacao.query.filter(PesquisaSatisfacao.nota_satisfacao.isnot(None)).count()
    pesquisas_puladas = PesquisaSatisfacao.query.filter_by(pulou=True).count()

    # Média de nota
    media_nota = db.session.query(func.avg(PesquisaSatisfacao.nota_satisfacao)).filter(
        PesquisaSatisfacao.nota_satisfacao.isnot(None)
    ).scalar() or 0
    media_nota = round(media_nota, 1)

    # Porcentagem de equipe atenciosa
    total_com_resposta_atenciosa = PesquisaSatisfacao.query.filter(PesquisaSatisfacao.equipe_atenciosa.isnot(None)).count()
    equipe_atenciosa_sim = PesquisaSatisfacao.query.filter_by(equipe_atenciosa=True).count()
    pct_atenciosa = round((equipe_atenciosa_sim / total_com_resposta_atenciosa * 100), 1) if total_com_resposta_atenciosa > 0 else 0

    # Distribuição de notas para gráfico
    distribuicao_notas = db.session.query(
        PesquisaSatisfacao.nota_satisfacao,
        func.count(PesquisaSatisfacao.id)
    ).filter(PesquisaSatisfacao.nota_satisfacao.isnot(None)
    ).group_by(PesquisaSatisfacao.nota_satisfacao
    ).order_by(PesquisaSatisfacao.nota_satisfacao
    ).all()

    notas_labels = [str(n[0]) for n in distribuicao_notas]
    notas_data = [n[1] for n in distribuicao_notas]

    # Comentários recentes
    comentarios_recentes = db.session.query(
        PesquisaSatisfacao,
        AgendamentoConsulta.paciente
    ).join(AgendamentoConsulta, AgendamentoConsulta.id == PesquisaSatisfacao.consulta_id
    ).filter(
        PesquisaSatisfacao.comentario.isnot(None),
        PesquisaSatisfacao.comentario != ''
    ).order_by(PesquisaSatisfacao.data_resposta.desc()
    ).limit(20).all()

    # =====================================================================
    # DADOS PARA GRÁFICOS - EVOLUÇÃO DIÁRIA (últimos 30 dias)
    # =====================================================================
    from datetime import timedelta
    data_inicio = datetime.utcnow() - timedelta(days=30)

    # MSG 1 (Disparos de agendamento) por dia
    msg1_por_dia = db.session.query(
        func.date(AgendamentoConsulta.data_envio_mensagem).label('dia'),
        func.count(AgendamentoConsulta.id).label('total')
    ).filter(
        AgendamentoConsulta.data_envio_mensagem >= data_inicio,
        AgendamentoConsulta.mensagem_enviada == True
    ).group_by(func.date(AgendamentoConsulta.data_envio_mensagem)
    ).order_by(func.date(AgendamentoConsulta.data_envio_mensagem)
    ).all()

    # Comprovantes enviados por dia (status CONFIRMADO)
    comprovantes_por_dia = db.session.query(
        func.date(AgendamentoConsulta.data_confirmacao).label('dia'),
        func.count(AgendamentoConsulta.id).label('total')
    ).filter(
        AgendamentoConsulta.data_confirmacao >= data_inicio,
        AgendamentoConsulta.status == 'CONFIRMADO'
    ).group_by(func.date(AgendamentoConsulta.data_confirmacao)
    ).order_by(func.date(AgendamentoConsulta.data_confirmacao)
    ).all()

    # Rejeitados por dia
    rejeitados_por_dia = db.session.query(
        func.date(AgendamentoConsulta.data_confirmacao).label('dia'),
        func.count(AgendamentoConsulta.id).label('total')
    ).filter(
        AgendamentoConsulta.data_confirmacao >= data_inicio,
        AgendamentoConsulta.status == 'REJEITADO'
    ).group_by(func.date(AgendamentoConsulta.data_confirmacao)
    ).order_by(func.date(AgendamentoConsulta.data_confirmacao)
    ).all()

    # Preparar dados para Chart.js
    dias_labels = [(data_inicio + timedelta(days=i)).strftime('%d/%m') for i in range(31)]

    # Mapear dados
    msg1_dict = {str(m.dia): m.total for m in msg1_por_dia}
    comprovantes_dict = {str(c.dia): c.total for c in comprovantes_por_dia}
    rejeitados_dict = {str(r.dia): r.total for r in rejeitados_por_dia}

    msg1_data = []
    comprovantes_data = []
    rejeitados_data = []

    for i in range(31):
        dia = (data_inicio + timedelta(days=i)).strftime('%Y-%m-%d')
        msg1_data.append(msg1_dict.get(dia, 0))
        comprovantes_data.append(comprovantes_dict.get(dia, 0))
        rejeitados_data.append(rejeitados_dict.get(dia, 0))

    # =====================================================================
    # DADOS PARA GRÁFICOS - MODO FILA (últimos 30 dias)
    # =====================================================================
    fila_enviados_por_dia = db.session.query(
        func.date(LogMsg.data).label('dia'),
        func.count(LogMsg.id).label('total')
    ).filter(
        LogMsg.data >= data_inicio,
        LogMsg.direcao == 'enviada'
    ).group_by(func.date(LogMsg.data)
    ).order_by(func.date(LogMsg.data)
    ).all()

    fila_confirmados_por_dia = db.session.query(
        func.date(Contato.data_resposta).label('dia'),
        func.count(Contato.id).label('total')
    ).filter(
        Contato.data_resposta >= data_inicio,
        Contato.confirmado == True
    ).group_by(func.date(Contato.data_resposta)
    ).order_by(func.date(Contato.data_resposta)
    ).all()

    fila_rejeitados_por_dia = db.session.query(
        func.date(Contato.data_resposta).label('dia'),
        func.count(Contato.id).label('total')
    ).filter(
        Contato.data_resposta >= data_inicio,
        Contato.rejeitado == True
    ).group_by(func.date(Contato.data_resposta)
    ).order_by(func.date(Contato.data_resposta)
    ).all()

    fila_dias_labels = [(data_inicio + timedelta(days=i)).strftime('%d/%m') for i in range(31)]

    fila_env_dict = {str(r.dia): r.total for r in fila_enviados_por_dia}
    fila_conf_dict = {str(r.dia): r.total for r in fila_confirmados_por_dia}
    fila_rej_dict = {str(r.dia): r.total for r in fila_rejeitados_por_dia}

    fila_enviados_data = []
    fila_confirmados_data = []
    fila_rejeitados_data = []
    for i in range(31):
        dia = (data_inicio + timedelta(days=i)).strftime('%Y-%m-%d')
        fila_enviados_data.append(fila_env_dict.get(dia, 0))
        fila_confirmados_data.append(fila_conf_dict.get(dia, 0))
        fila_rejeitados_data.append(fila_rej_dict.get(dia, 0))

    # Distribuição de status atual (Fila)
    fila_status_counts = db.session.query(
        Contato.status,
        func.count(Contato.id).label('total')
    ).group_by(Contato.status).all()
    fila_status_labels = [r.status or 'pendente' for r in fila_status_counts]
    fila_status_data = [r.total for r in fila_status_counts]

    # =====================================================================
    # ESPECIALIDADES MAIS FREQUENTES
    # =====================================================================
    especialidades_top = db.session.query(
        AgendamentoConsulta.especialidade,
        func.count(AgendamentoConsulta.id).label('total'),
        func.sum(case((AgendamentoConsulta.status == 'CONFIRMADO', 1), else_=0)).label('confirmados'),
        func.sum(case((AgendamentoConsulta.status == 'REJEITADO', 1), else_=0)).label('rejeitados')
    ).filter(AgendamentoConsulta.especialidade.isnot(None)
    ).group_by(AgendamentoConsulta.especialidade
    ).order_by(func.count(AgendamentoConsulta.id).desc()
    ).limit(10).all()

    # =====================================================================
    # NOTAS MÉDIAS POR ESPECIALIDADE (Pesquisa de Satisfação)
    # =====================================================================
    notas_por_especialidade = db.session.query(
        PesquisaSatisfacao.especialidade,
        func.avg(PesquisaSatisfacao.nota_satisfacao).label('media_nota'),
        func.count(PesquisaSatisfacao.id).label('total_respostas'),
        func.sum(case((PesquisaSatisfacao.equipe_atenciosa == True, 1), else_=0)).label('atenciosa_sim'),
        func.count(case((PesquisaSatisfacao.equipe_atenciosa.isnot(None), 1))).label('total_atenciosa')
    ).filter(
        PesquisaSatisfacao.especialidade.isnot(None),
        PesquisaSatisfacao.nota_satisfacao.isnot(None)
    ).group_by(PesquisaSatisfacao.especialidade
    ).order_by(func.count(PesquisaSatisfacao.id).desc()
    ).limit(20).all()

    # =====================================================================
    # ESTATÍSTICAS DETALHADAS POR USUÁRIO
    # =====================================================================
    # Performance por usuário com mais detalhes
    stats_query = db.session.query(
        Usuario.id,
        Usuario.nome,
        Usuario.tipo_sistema,
        # Campanhas criadas
        func.count(func.distinct(CampanhaConsulta.id)).label('campanhas_criadas'),
        # Agendamentos
        func.count(AgendamentoConsulta.id).label('total_agendamentos'),
        func.sum(case((AgendamentoConsulta.status == 'CONFIRMADO', 1), else_=0)).label('confirmados'),
        func.sum(case((AgendamentoConsulta.status == 'REJEITADO', 1), else_=0)).label('rejeitados'),
        func.sum(case((AgendamentoConsulta.mensagem_enviada == True, 1), else_=0)).label('enviados'),
        # Pesquisas de satisfação
        func.count(PesquisaSatisfacao.id).label('total_pesquisas'),
        func.avg(PesquisaSatisfacao.nota_satisfacao).label('media_nota_usuario')
    ).outerjoin(CampanhaConsulta, CampanhaConsulta.criador_id == Usuario.id
    ).outerjoin(AgendamentoConsulta, AgendamentoConsulta.campanha_id == CampanhaConsulta.id
    ).outerjoin(PesquisaSatisfacao, PesquisaSatisfacao.consulta_id == AgendamentoConsulta.id
    )

    # Aplicar filtro de modo
    if modo_selecionado == 'consulta':
        stats_query = stats_query.filter(Usuario.tipo_sistema == 'AGENDAMENTO_CONSULTA')
    elif modo_selecionado == 'fila':
        stats_query = stats_query.filter(Usuario.tipo_sistema == 'BUSCA_ATIVA')

    stats_detalhadas_usuario = stats_query.group_by(Usuario.id, Usuario.nome, Usuario.tipo_sistema
    ).having(func.count(AgendamentoConsulta.id) > 0
    ).order_by(func.count(AgendamentoConsulta.id).desc()
    ).limit(20).all()

    # Total de comentários disponíveis
    total_comentarios = db.session.query(func.count(PesquisaSatisfacao.id)).filter(
        PesquisaSatisfacao.comentario.isnot(None),
        PesquisaSatisfacao.comentario != ''
    ).scalar() or 0

    # =====================================================================
    # TOTAIS GERAIS DO SISTEMA
    # =====================================================================
    total_msgs_consulta = LogMsgConsulta.query.count()
    total_msgs_fila = LogMsg.query.count()
    total_msgs_sistema = total_msgs_consulta + total_msgs_fila

    total_campanhas = total_campanhas_fila + total_campanhas_consulta
    total_enviados = fila_total_enviados + consulta_enviados
    total_confirmados = fila_total_confirmados + consulta_confirmados

    taxa_confirmacao_geral = round((total_confirmados / total_enviados * 100), 1) if total_enviados > 0 else 0

    return render_template('admin_dashboard.html',
        # Modo selecionado
        modo_selecionado=modo_selecionado,

        # Usuários
        total_usuarios=total_usuarios,
        usuarios_ativos=usuarios_ativos,
        usuarios_admin=usuarios_admin,
        usuarios_fila=usuarios_fila,
        usuarios_consulta=usuarios_consulta,

        # Modo Fila
        total_campanhas_fila=total_campanhas_fila,
        campanhas_fila_ativas=campanhas_fila_ativas,
        fila_total_contatos=fila_total_contatos,
        fila_total_enviados=fila_total_enviados,
        fila_total_confirmados=fila_total_confirmados,
        fila_total_rejeitados=fila_total_rejeitados,
        fila_total_erros=fila_total_erros,
        fila_taxa_confirmacao=fila_taxa_confirmacao,

        # Modo Consulta
        total_campanhas_consulta=total_campanhas_consulta,
        campanhas_consulta_ativas=campanhas_consulta_ativas,
        consulta_total=consulta_total,
        consulta_enviados=consulta_enviados,
        consulta_confirmados=consulta_confirmados,
        consulta_rejeitados=consulta_rejeitados,
        consulta_taxa_confirmacao=consulta_taxa_confirmacao,

        # Estatísticas detalhadas de agendamento
        total_disparos_msg1=total_disparos_msg1,
        total_comprovantes_enviados=total_comprovantes_enviados,
        total_aguardando_comprovante=total_aguardando_comprovante,
        total_cancelados=total_cancelados,
        desempenho_usuarios=desempenho_usuarios,

        # Mensagens por usuário
        msgs_por_usuario_consulta=msgs_por_usuario_consulta,
        msgs_por_usuario_fila=msgs_por_usuario_fila,
        confirmacao_por_usuario=confirmacao_por_usuario,
        stats_detalhadas_usuario=stats_detalhadas_usuario,

        # Pesquisas de satisfação
        total_pesquisas=total_pesquisas,
        pesquisas_respondidas=pesquisas_respondidas,
        pesquisas_puladas=pesquisas_puladas,
        media_nota=media_nota,
        pct_atenciosa=pct_atenciosa,
        notas_labels=notas_labels,
        notas_data=notas_data,
        comentarios_recentes=comentarios_recentes,
        total_comentarios=total_comentarios,

        # Dados para gráficos - Modo Consulta
        dias_labels=dias_labels,
        msg1_data=msg1_data,
        comprovantes_data=comprovantes_data,
        rejeitados_data=rejeitados_data,

        # Dados para gráficos - Modo Fila
        fila_dias_labels=fila_dias_labels,
        fila_enviados_data=fila_enviados_data,
        fila_confirmados_data=fila_confirmados_data,
        fila_rejeitados_data=fila_rejeitados_data,
        fila_status_labels=fila_status_labels,
        fila_status_data=fila_status_data,

        # Especialidades
        especialidades_top=especialidades_top,
        notas_por_especialidade=notas_por_especialidade,

        # Totais gerais
        total_msgs_sistema=total_msgs_sistema,
        total_campanhas=total_campanhas,
        total_enviados=total_enviados,
        total_confirmados=total_confirmados,
        taxa_confirmacao_geral=taxa_confirmacao_geral,

        # Lista de usuários para exportação
        usuarios_export=Usuario.query.filter_by(tipo_sistema='AGENDAMENTO_CONSULTA').order_by(Usuario.nome).all(),
        usuarios_fila_export=Usuario.query.filter_by(tipo_sistema='BUSCA_ATIVA').order_by(Usuario.nome).all()
    )
@bp.route('/admin/exportar', methods=['GET', 'POST'])
@login_required
@admin_required
def exportar_dados():
    """
    Exportar dados de agendamentos para Excel com filtros.
    Permite filtrar por status, tipo, especialidade, usuário, datas, etc.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    # Obter parâmetros de filtro (GET ou POST)
    if request.method == 'POST':
        filtros = request.form
    else:
        filtros = request.args

    # Filtros disponíveis
    status_filtro = filtros.getlist('status') or filtros.get('status', '').split(',')
    status_filtro = [s.strip() for s in status_filtro if s.strip()]

    tipo_filtro = filtros.getlist('tipo') or filtros.get('tipo', '').split(',')
    tipo_filtro = [t.strip() for t in tipo_filtro if t.strip()]

    usuario_filtro = filtros.get('usuario_id', '')
    especialidade_filtro = filtros.get('especialidade', '')
    data_inicio = filtros.get('data_inicio', '')
    data_fim = filtros.get('data_fim', '')

    # Filtros de telefone
    incluir_telefones_validos = filtros.get('telefones_validos', 'true').lower() == 'true'
    incluir_telefones_invalidos = filtros.get('telefones_invalidos', 'true').lower() == 'true'

    # Incluir pesquisa de satisfação
    incluir_pesquisa = filtros.get('incluir_pesquisa', 'true').lower() == 'true'

    # Query base
    query = AgendamentoConsulta.query

    # Aplicar filtros
    if status_filtro:
        query = query.filter(AgendamentoConsulta.status.in_(status_filtro))

    if tipo_filtro:
        query = query.filter(AgendamentoConsulta.tipo.in_(tipo_filtro))

    if usuario_filtro:
        query = query.filter(AgendamentoConsulta.usuario_id == int(usuario_filtro))

    if especialidade_filtro:
        query = query.filter(AgendamentoConsulta.especialidade.ilike(f'%{especialidade_filtro}%'))

    if data_inicio:
        try:
            dt_inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
            query = query.filter(AgendamentoConsulta.created_at >= dt_inicio)
        except:
            pass

    if data_fim:
        try:
            dt_fim = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(AgendamentoConsulta.created_at < dt_fim)
        except:
            pass

    # Ordenar por data de criação (mais recente primeiro)
    agendamentos = query.order_by(AgendamentoConsulta.created_at.desc()).all()

    # Criar workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Agendamentos"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Cabeçalhos
    headers = [
        "ID", "Campanha", "Usuário", "Paciente", "Código AGHU", "Cod Master",
        "Tipo", "Especialidade", "Sub-Especialidade", "Procedência",
        "Médico Solicitante", "Data AGHU", "Grade AGHU", "Prioridade",
        "Telefone Cadastro", "Telefone Registro", "Telefones Válidos", "Telefones Inválidos",
        "Status", "Motivo Rejeição", "Data Criação", "Data Envio MSG1",
        "Data Confirmação", "Data Rejeição", "Telefone que Confirmou",
        "Comprovante Enviado", "Tentativas de Contato",
        "Nova Data (Reagend.)", "Nova Hora (Reagend.)",
        "Observações", "Exames"
    ]

    if incluir_pesquisa:
        headers.extend(["Nota Satisfação", "Equipe Atenciosa", "Comentário Pesquisa"])

    # Escrever cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Escrever dados
    row_num = 2
    for ag in agendamentos:
        # Coletar telefones válidos e inválidos
        telefones_validos = []
        telefones_invalidos = []
        for tel in ag.telefones:
            if tel.invalido:
                telefones_invalidos.append(tel.numero)
            else:
                telefones_validos.append(tel.numero)

        # Aplicar filtro de telefones
        if not incluir_telefones_validos and not incluir_telefones_invalidos:
            continue
        if not incluir_telefones_validos and len(telefones_validos) > 0 and len(telefones_invalidos) == 0:
            continue
        if not incluir_telefones_invalidos and len(telefones_invalidos) > 0 and len(telefones_validos) == 0:
            continue

        # Buscar pesquisa de satisfação
        pesquisa = None
        if incluir_pesquisa:
            pesquisa = PesquisaSatisfacao.query.filter_by(consulta_id=ag.id).first()

        # Dados da linha
        row_data = [
            ag.id,
            ag.campanha.nome if ag.campanha else '',
            ag.usuario.nome if ag.usuario else '',
            ag.paciente,
            ag.codigo_aghu or '',
            ag.cod_master or '',
            ag.tipo or '',
            ag.especialidade or '',
            ag.sub_especialidade or '',
            ag.procedencia or '',
            ag.medico_solicitante or '',
            ag.data_aghu or '',
            ag.grade_aghu or '',
            ag.prioridade or '',
            ag.telefone_cadastro or '',
            ag.telefone_registro or '',
            ', '.join(telefones_validos),
            ', '.join(telefones_invalidos),
            ag.status or '',
            ag.motivo_rejeicao or '',
            ag.created_at.strftime('%d/%m/%Y %H:%M') if ag.created_at else '',
            ag.data_envio_mensagem.strftime('%d/%m/%Y %H:%M') if ag.data_envio_mensagem else '',
            ag.data_confirmacao.strftime('%d/%m/%Y %H:%M') if ag.data_confirmacao else '',
            ag.data_rejeicao.strftime('%d/%m/%Y %H:%M') if ag.data_rejeicao else '',
            ag.telefone_confirmacao or '',
            'Sim' if ag.comprovante_path else 'Não',
            ag.tentativas_contato or 0,
            ag.nova_data or '',
            ag.nova_hora or '',
            ag.observacoes or '',
            ag.exames or ''
        ]

        if incluir_pesquisa:
            if pesquisa:
                row_data.extend([
                    pesquisa.nota_satisfacao if pesquisa.nota_satisfacao else ('Pulou' if pesquisa.pulou else ''),
                    'Sim' if pesquisa.equipe_atenciosa else ('Não' if pesquisa.equipe_atenciosa is False else ''),
                    pesquisa.comentario or ''
                ])
            else:
                row_data.extend(['', '', ''])

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        row_num += 1

    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Colunas específicas mais largas
    ws.column_dimensions['D'].width = 30  # Paciente
    ws.column_dimensions['H'].width = 25  # Especialidade
    ws.column_dimensions['Q'].width = 25  # Telefones Válidos
    ws.column_dimensions['R'].width = 25  # Telefones Inválidos
    ws.column_dimensions['T'].width = 40  # Motivo Rejeição
    ws.column_dimensions['AD'].width = 40  # Observações
    ws.column_dimensions['AE'].width = 40  # Exames
    if incluir_pesquisa:
        ws.column_dimensions['AH'].width = 40  # Comentário Pesquisa

    # Congelar primeira linha (cabeçalhos)
    ws.freeze_panes = 'A2'

    # Salvar em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Nome do arquivo com data
    filename = f"agendamentos_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
@bp.route('/admin/exportar/fila', methods=['GET', 'POST'])
@login_required
@admin_required
def exportar_fila():
    """
    Exportar dados de campanhas de Fila (Busca Ativa) para Excel com filtros.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from io import BytesIO

    # Obter parâmetros de filtro
    if request.method == 'POST':
        filtros = request.form
    else:
        filtros = request.args

    # Filtros
    status_filtro = filtros.getlist('status_fila') or []
    usuario_filtro = filtros.get('usuario_id_fila', '')
    procedimento_filtro = filtros.get('procedimento_fila', '')
    data_inicio = filtros.get('data_inicio_fila', '')
    data_fim = filtros.get('data_fim_fila', '')

    # Query base - Contatos das campanhas
    query = Contato.query.join(Campanha)

    # Aplicar filtros
    if status_filtro:
        # Mapear status para campos booleanos
        conditions = []
        if 'confirmado' in status_filtro:
            conditions.append(Contato.confirmado == True)
        if 'rejeitado' in status_filtro:
            conditions.append(Contato.rejeitado == True)
        if 'enviado' in status_filtro:
            conditions.append((Contato.enviado == True) & (Contato.confirmado == False) & (Contato.rejeitado == False))
        if 'aguardando' in status_filtro:
            conditions.append(Contato.enviado == False)
        if 'erro' in status_filtro:
            conditions.append(Contato.erro != None)
        if conditions:
            from sqlalchemy import or_
            query = query.filter(or_(*conditions))

    if usuario_filtro:
        query = query.filter(Campanha.usuario_id == int(usuario_filtro))

    if procedimento_filtro:
        query = query.filter(Contato.procedimento.ilike(f'%{procedimento_filtro}%'))

    if data_inicio:
        try:
            dt_inicio = datetime.strptime(data_inicio, '%Y-%m-%d')
            query = query.filter(Contato.data_criacao >= dt_inicio)
        except:
            pass

    if data_fim:
        try:
            dt_fim = datetime.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(Contato.data_criacao < dt_fim)
        except:
            pass

    # Ordenar
    contatos = query.order_by(Contato.data_criacao.desc()).all()

    # Criar workbook Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fila Cirurgica"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="343A40", end_color="343A40", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Cabeçalhos
    headers = [
        "ID", "Campanha", "Usuário", "Nome", "Data Nascimento", "Telefones",
        "Procedimento", "Status", "Enviado", "Data Envio", "Confirmado",
        "Data Confirmação", "Rejeitado", "Data Rejeição", "Resposta",
        "Erro", "Data Criação"
    ]

    # Escrever cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Escrever dados
    row_num = 2
    for c in contatos:
        # Determinar status
        if c.confirmado:
            status = 'CONFIRMADO'
        elif c.rejeitado:
            status = 'REJEITADO'
        elif c.erro:
            status = 'ERRO'
        elif c.enviado:
            status = 'ENVIADO'
        else:
            status = 'AGUARDANDO'

        # Dados de envio derivados dos telefones
        tel_enviado = next((t for t in c.telefones if t.enviado), None)
        data_envio = tel_enviado.data_envio if tel_enviado else None
        data_confirmacao = c.data_resposta if c.confirmado else None

        row_data = [
            c.id,
            c.campanha.nome if c.campanha else '',
            c.campanha.usuario.nome if c.campanha and c.campanha.usuario else '',
            c.nome,
            c.data_nascimento.strftime('%d/%m/%Y') if c.data_nascimento else '',
            c.telefones_str(),
            c.procedimento or '',
            status,
            'Sim' if tel_enviado else 'Não',
            data_envio.strftime('%d/%m/%Y %H:%M') if data_envio else '',
            'Sim' if c.confirmado else 'Não',
            data_confirmacao.strftime('%d/%m/%Y %H:%M') if data_confirmacao else '',
            'Sim' if c.rejeitado else 'Não',
            c.data_rejeicao.strftime('%d/%m/%Y %H:%M') if c.data_rejeicao else '',
            c.resposta or '',
            c.erro or '',
            c.data_criacao.strftime('%d/%m/%Y %H:%M') if c.data_criacao else ''
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        row_num += 1

    # Ajustar largura das colunas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Colunas específicas
    ws.column_dimensions['D'].width = 30  # Nome
    ws.column_dimensions['F'].width = 25  # Telefones
    ws.column_dimensions['G'].width = 30  # Procedimento
    ws.column_dimensions['O'].width = 40  # Resposta
    ws.column_dimensions['P'].width = 30  # Erro

    # Congelar primeira linha
    ws.freeze_panes = 'A2'

    # Salvar
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"fila_cirurgica_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
@bp.route('/admin/usuario/<int:usuario_id>')
@login_required
@admin_required
def usuario_detalhes(usuario_id):
    """Página de detalhes completos de um usuário específico"""
    from sqlalchemy import func, case
    from datetime import timedelta

    usuario = Usuario.query.get_or_404(usuario_id)

    # =====================================================================
    # ESTATÍSTICAS GERAIS DO USUÁRIO (MODO CONSULTA)
    # =====================================================================
    # Campanhas criadas pelo usuário
    campanhas_usuario = CampanhaConsulta.query.filter_by(criador_id=usuario_id).all()
    total_campanhas = len(campanhas_usuario)
    campanhas_ativas = sum(1 for c in campanhas_usuario if c.status in ['pronta', 'enviando'])

    # IDs das campanhas do usuário para queries
    campanha_ids = [c.id for c in campanhas_usuario]

    if campanha_ids:
        # Estatísticas de agendamentos
        stats_agendamentos = db.session.query(
            func.count(AgendamentoConsulta.id).label('total'),
            func.sum(case((AgendamentoConsulta.mensagem_enviada == True, 1), else_=0)).label('msg1_enviadas'),
            func.sum(case((AgendamentoConsulta.status == 'CONFIRMADO', 1), else_=0)).label('confirmados'),
            func.sum(case((AgendamentoConsulta.status == 'REJEITADO', 1), else_=0)).label('rejeitados'),
            func.sum(case((AgendamentoConsulta.status == 'AGUARDANDO_COMPROVANTE', 1), else_=0)).label('aguardando'),
            func.sum(case((AgendamentoConsulta.status == 'CANCELADO', 1), else_=0)).label('cancelados')
        ).filter(AgendamentoConsulta.campanha_id.in_(campanha_ids)).first()

        total_agendamentos = stats_agendamentos.total or 0
        msg1_enviadas = stats_agendamentos.msg1_enviadas or 0
        confirmados = stats_agendamentos.confirmados or 0
        rejeitados = stats_agendamentos.rejeitados or 0
        aguardando = stats_agendamentos.aguardando or 0
        cancelados = stats_agendamentos.cancelados or 0

        # Taxa de sucesso (comprovantes enviados / MSG 1 enviadas)
        taxa_sucesso = round((confirmados / msg1_enviadas * 100), 1) if msg1_enviadas > 0 else 0

        # =====================================================================
        # ATIVIDADE DOS ÚLTIMOS 30 DIAS
        # =====================================================================
        data_inicio = datetime.utcnow() - timedelta(days=30)

        # MSG 1 enviadas por dia
        msg1_por_dia = db.session.query(
            func.date(AgendamentoConsulta.data_envio_mensagem).label('dia'),
            func.count(AgendamentoConsulta.id).label('total')
        ).filter(
            AgendamentoConsulta.campanha_id.in_(campanha_ids),
            AgendamentoConsulta.data_envio_mensagem >= data_inicio,
            AgendamentoConsulta.mensagem_enviada == True
        ).group_by(func.date(AgendamentoConsulta.data_envio_mensagem)
        ).order_by(func.date(AgendamentoConsulta.data_envio_mensagem)
        ).all()

        # Comprovantes enviados por dia
        comprovantes_por_dia = db.session.query(
            func.date(AgendamentoConsulta.data_confirmacao).label('dia'),
            func.count(AgendamentoConsulta.id).label('total')
        ).filter(
            AgendamentoConsulta.campanha_id.in_(campanha_ids),
            AgendamentoConsulta.data_confirmacao >= data_inicio,
            AgendamentoConsulta.status == 'CONFIRMADO'
        ).group_by(func.date(AgendamentoConsulta.data_confirmacao)
        ).order_by(func.date(AgendamentoConsulta.data_confirmacao)
        ).all()

        # Rejeitados por dia
        rejeitados_por_dia = db.session.query(
            func.date(AgendamentoConsulta.data_confirmacao).label('dia'),
            func.count(AgendamentoConsulta.id).label('total')
        ).filter(
            AgendamentoConsulta.campanha_id.in_(campanha_ids),
            AgendamentoConsulta.data_confirmacao >= data_inicio,
            AgendamentoConsulta.status == 'REJEITADO'
        ).group_by(func.date(AgendamentoConsulta.data_confirmacao)
        ).order_by(func.date(AgendamentoConsulta.data_confirmacao)
        ).all()

        # Preparar dados para Chart.js
        dias_labels = [(data_inicio + timedelta(days=i)).strftime('%d/%m') for i in range(31)]
        msg1_dict = {str(m.dia): m.total for m in msg1_por_dia}
        comprovantes_dict = {str(c.dia): c.total for c in comprovantes_por_dia}
        rejeitados_dict = {str(r.dia): r.total for r in rejeitados_por_dia}

        msg1_data = []
        comprovantes_data = []
        rejeitados_data = []

        for i in range(31):
            dia = (data_inicio + timedelta(days=i)).strftime('%Y-%m-%d')
            msg1_data.append(msg1_dict.get(dia, 0))
            comprovantes_data.append(comprovantes_dict.get(dia, 0))
            rejeitados_data.append(rejeitados_dict.get(dia, 0))

        # =====================================================================
        # ATIVIDADE POR HORA DO DIA (padrão de trabalho)
        # =====================================================================
        atividade_por_hora = db.session.query(
            func.extract('hour', AgendamentoConsulta.data_envio_mensagem).label('hora'),
            func.count(AgendamentoConsulta.id).label('total')
        ).filter(
            AgendamentoConsulta.campanha_id.in_(campanha_ids),
            AgendamentoConsulta.mensagem_enviada == True
        ).group_by(func.extract('hour', AgendamentoConsulta.data_envio_mensagem)
        ).order_by(func.extract('hour', AgendamentoConsulta.data_envio_mensagem)
        ).all()

        horas_labels = [f"{h:02d}:00" for h in range(24)]
        horas_dict = {int(a.hora): a.total for a in atividade_por_hora if a.hora is not None}
        horas_data = [horas_dict.get(h, 0) for h in range(24)]

        # =====================================================================
        # ESPECIALIDADES MAIS FREQUENTES DO USUÁRIO
        # =====================================================================
        especialidades_usuario = db.session.query(
            AgendamentoConsulta.especialidade,
            func.count(AgendamentoConsulta.id).label('total'),
            func.sum(case((AgendamentoConsulta.status == 'CONFIRMADO', 1), else_=0)).label('confirmados'),
            func.sum(case((AgendamentoConsulta.status == 'REJEITADO', 1), else_=0)).label('rejeitados')
        ).filter(
            AgendamentoConsulta.campanha_id.in_(campanha_ids),
            AgendamentoConsulta.especialidade.isnot(None)
        ).group_by(AgendamentoConsulta.especialidade
        ).order_by(func.count(AgendamentoConsulta.id).desc()
        ).limit(10).all()

        # =====================================================================
        # PESQUISAS DE SATISFAÇÃO DO USUÁRIO
        # =====================================================================
        # Buscar IDs dos agendamentos via query (não via relação)
        agendamento_ids = [a.id for a in AgendamentoConsulta.query.filter(
            AgendamentoConsulta.campanha_id.in_(campanha_ids)
        ).all()]

        if agendamento_ids:
            pesquisas_stats = db.session.query(
                func.count(PesquisaSatisfacao.id).label('total'),
                func.count(case((PesquisaSatisfacao.nota_satisfacao.isnot(None), 1))).label('respondidas'),
                func.avg(PesquisaSatisfacao.nota_satisfacao).label('media_nota'),
                func.sum(case((PesquisaSatisfacao.equipe_atenciosa == True, 1), else_=0)).label('atenciosa_sim'),
                func.count(case((PesquisaSatisfacao.equipe_atenciosa.isnot(None), 1))).label('total_atenciosa')
            ).filter(PesquisaSatisfacao.consulta_id.in_(agendamento_ids)).first()

            total_pesquisas = pesquisas_stats.total or 0
            pesquisas_respondidas = pesquisas_stats.respondidas or 0
            media_nota = round(pesquisas_stats.media_nota or 0, 1)
            pct_atenciosa = round((pesquisas_stats.atenciosa_sim or 0) / pesquisas_stats.total_atenciosa * 100, 1) if pesquisas_stats.total_atenciosa else 0

            # Distribuição de notas do usuário
            distribuicao_notas = db.session.query(
                PesquisaSatisfacao.nota_satisfacao,
                func.count(PesquisaSatisfacao.id)
            ).filter(
                PesquisaSatisfacao.consulta_id.in_(agendamento_ids),
                PesquisaSatisfacao.nota_satisfacao.isnot(None)
            ).group_by(PesquisaSatisfacao.nota_satisfacao
            ).order_by(PesquisaSatisfacao.nota_satisfacao
            ).all()

            notas_labels = [str(n[0]) for n in distribuicao_notas]
            notas_data = [n[1] for n in distribuicao_notas]

            # Comentários recentes do usuário
            comentarios_usuario = db.session.query(
                PesquisaSatisfacao,
                AgendamentoConsulta.paciente
            ).join(AgendamentoConsulta, AgendamentoConsulta.id == PesquisaSatisfacao.consulta_id
            ).filter(
                PesquisaSatisfacao.consulta_id.in_(agendamento_ids),
                PesquisaSatisfacao.comentario.isnot(None),
                PesquisaSatisfacao.comentario != ''
            ).order_by(PesquisaSatisfacao.data_resposta.desc()
            ).limit(10).all()
        else:
            total_pesquisas = 0
            pesquisas_respondidas = 0
            media_nota = 0
            pct_atenciosa = 0
            notas_labels = []
            notas_data = []
            comentarios_usuario = []

        # =====================================================================
        # CAMPANHAS RECENTES DO USUÁRIO
        # =====================================================================
        campanhas_recentes = db.session.query(
            CampanhaConsulta,
            func.count(AgendamentoConsulta.id).label('total_agendamentos'),
            func.sum(case((AgendamentoConsulta.mensagem_enviada == True, 1), else_=0)).label('msg1_enviadas'),
            func.sum(case((AgendamentoConsulta.status == 'CONFIRMADO', 1), else_=0)).label('confirmados')
        ).outerjoin(AgendamentoConsulta, AgendamentoConsulta.campanha_id == CampanhaConsulta.id
        ).filter(CampanhaConsulta.criador_id == usuario_id
        ).group_by(CampanhaConsulta.id
        ).order_by(CampanhaConsulta.data_criacao.desc()
        ).limit(10).all()

        # =====================================================================
        # MÉTRICAS DE TEMPO
        # =====================================================================
        # Tempo médio entre envio MSG1 e confirmação (apenas confirmados)
        tempo_medio_confirmacao = db.session.query(
            func.avg(
                func.extract('epoch', AgendamentoConsulta.data_confirmacao) -
                func.extract('epoch', AgendamentoConsulta.data_envio_mensagem)
            )
        ).filter(
            AgendamentoConsulta.campanha_id.in_(campanha_ids),
            AgendamentoConsulta.status == 'CONFIRMADO',
            AgendamentoConsulta.data_envio_mensagem.isnot(None),
            AgendamentoConsulta.data_confirmacao.isnot(None)
        ).scalar()

        if tempo_medio_confirmacao:
            horas_media = int(tempo_medio_confirmacao // 3600)
            minutos_media = int((tempo_medio_confirmacao % 3600) // 60)
            tempo_medio_str = f"{horas_media}h {minutos_media}min"
        else:
            tempo_medio_str = "N/A"

    else:
        # Usuário sem campanhas
        total_agendamentos = 0
        msg1_enviadas = 0
        confirmados = 0
        rejeitados = 0
        aguardando = 0
        cancelados = 0
        taxa_sucesso = 0
        dias_labels = []
        msg1_data = []
        comprovantes_data = []
        rejeitados_data = []
        horas_labels = []
        horas_data = []
        especialidades_usuario = []
        total_pesquisas = 0
        pesquisas_respondidas = 0
        media_nota = 0
        pct_atenciosa = 0
        notas_labels = []
        notas_data = []
        comentarios_usuario = []
        campanhas_recentes = []
        tempo_medio_str = "N/A"

    return render_template('admin_usuario_detalhes.html',
        usuario=usuario,
        total_campanhas=total_campanhas,
        campanhas_ativas=campanhas_ativas,
        total_agendamentos=total_agendamentos,
        msg1_enviadas=msg1_enviadas,
        confirmados=confirmados,
        rejeitados=rejeitados,
        aguardando=aguardando,
        cancelados=cancelados,
        taxa_sucesso=taxa_sucesso,
        dias_labels=dias_labels,
        msg1_data=msg1_data,
        comprovantes_data=comprovantes_data,
        rejeitados_data=rejeitados_data,
        horas_labels=horas_labels,
        horas_data=horas_data,
        especialidades_usuario=especialidades_usuario,
        total_pesquisas=total_pesquisas,
        pesquisas_respondidas=pesquisas_respondidas,
        media_nota=media_nota,
        pct_atenciosa=pct_atenciosa,
        notas_labels=notas_labels,
        notas_data=notas_data,
        comentarios_usuario=comentarios_usuario,
        campanhas_recentes=campanhas_recentes,
        tempo_medio_str=tempo_medio_str
    )
@bp.route('/admin/usuario/<int:usuario_id>/deletar', methods=['POST'])
@login_required
@admin_required
def deletar_usuario(usuario_id):
    """Deletar um usuário do sistema"""
    usuario = Usuario.query.get_or_404(usuario_id)

    # Não permitir deletar o próprio usuário logado
    if usuario.id == current_user.id:
        flash('Voce nao pode deletar sua propria conta!', 'danger')
        return redirect(url_for('admin.usuario_detalhes', usuario_id=usuario_id))

    # Não permitir deletar outros admins (segurança)
    if usuario.is_admin and not current_user.is_admin:
        flash('Apenas administradores podem deletar outros administradores!', 'danger')
        return redirect(url_for('admin.usuario_detalhes', usuario_id=usuario_id))

    nome_usuario = usuario.nome

    try:
        # Deletar campanhas de consulta do usuário e seus agendamentos
        campanhas_consulta = CampanhaConsulta.query.filter_by(criador_id=usuario_id).all()
        for campanha in campanhas_consulta:
            # Deletar pesquisas de satisfação dos agendamentos
            for agendamento in campanha.agendamentos:
                PesquisaSatisfacao.query.filter_by(consulta_id=agendamento.id).delete()
            # Deletar agendamentos
            AgendamentoConsulta.query.filter_by(campanha_id=campanha.id).delete()
            # Deletar logs de mensagens
            LogMsgConsulta.query.filter_by(campanha_id=campanha.id).delete()
        # Deletar campanhas de consulta
        CampanhaConsulta.query.filter_by(criador_id=usuario_id).delete()

        # Deletar campanhas de fila do usuário
        campanhas_fila = Campanha.query.filter_by(criador_id=usuario_id).all()
        for campanha in campanhas_fila:
            # Deletar contatos
            Contato.query.filter_by(campanha_id=campanha.id).delete()
            # Deletar logs de mensagens
            LogMsg.query.filter_by(campanha_id=campanha.id).delete()
        # Deletar campanhas de fila
        Campanha.query.filter_by(criador_id=usuario_id).delete()

        # Deletar configurações do WhatsApp do usuário
        ConfigWhatsApp.query.filter_by(usuario_id=usuario_id).delete()

        # Deletar o usuário
        db.session.delete(usuario)
        db.session.commit()

        flash(f'Usuario "{nome_usuario}" deletado com sucesso!', 'success')
        return redirect(url_for('admin.dashboard'))

    except Exception as e:
        db.session.rollback()
        flash(f'Erro ao deletar usuario: {str(e)}', 'danger')
        return redirect(url_for('admin.usuario_detalhes', usuario_id=usuario_id))
@bp.route('/admin/comentarios')
@login_required
@admin_required
def comentarios():
    """Página para visualizar todos os comentários da pesquisa de satisfação"""
    from sqlalchemy import func, case

    # Parâmetros de filtro da query string
    especialidade_filtro = request.args.get('especialidade')
    nota_min = request.args.get('nota_min', type=int)
    nota_max = request.args.get('nota_max', type=int)
    usuario_filtro = request.args.get('usuario')
    data_inicio = request.args.get('data_inicio')
    data_fim = request.args.get('data_fim')
    busca = request.args.get('busca', '').strip()
    pagina = request.args.get('pagina', 1, type=int)
    por_pagina = 50

    # Query base
    query = db.session.query(
        PesquisaSatisfacao,
        AgendamentoConsulta.paciente,
        Usuario.nome.label('usuario_nome')
    ).join(AgendamentoConsulta, AgendamentoConsulta.id == PesquisaSatisfacao.consulta_id
    ).join(CampanhaConsulta, CampanhaConsulta.id == AgendamentoConsulta.campanha_id
    ).join(Usuario, Usuario.id == CampanhaConsulta.criador_id
    ).filter(
        PesquisaSatisfacao.comentario.isnot(None),
        PesquisaSatisfacao.comentario != ''
    )

    # Aplicar filtros
    if especialidade_filtro:
        query = query.filter(PesquisaSatisfacao.especialidade == especialidade_filtro)

    if nota_min is not None:
        query = query.filter(PesquisaSatisfacao.nota_satisfacao >= nota_min)

    if nota_max is not None:
        query = query.filter(PesquisaSatisfacao.nota_satisfacao <= nota_max)

    if usuario_filtro:
        query = query.filter(Usuario.nome.ilike(f'%{usuario_filtro}%'))

    if data_inicio:
        try:
            from datetime import datetime as dt
            data_inicio_dt = dt.strptime(data_inicio, '%Y-%m-%d')
            query = query.filter(PesquisaSatisfacao.data_resposta >= data_inicio_dt)
        except:
            pass

    if data_fim:
        try:
            from datetime import datetime as dt, timedelta
            data_fim_dt = dt.strptime(data_fim, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(PesquisaSatisfacao.data_resposta < data_fim_dt)
        except:
            pass

    if busca:
        query = query.filter(
            db.or_(
                PesquisaSatisfacao.comentario.ilike(f'%{busca}%'),
                AgendamentoConsulta.paciente.ilike(f'%{busca}%')
            )
        )

    # Ordenar por data mais recente
    query = query.order_by(PesquisaSatisfacao.data_resposta.desc())

    # Paginação
    total_registros = query.count()
    total_paginas = (total_registros + por_pagina - 1) // por_pagina
    comentarios = query.limit(por_pagina).offset((pagina - 1) * por_pagina).all()

    # Listas para filtros
    especialidades_disponiveis = db.session.query(
        PesquisaSatisfacao.especialidade
    ).filter(
        PesquisaSatisfacao.especialidade.isnot(None),
        PesquisaSatisfacao.comentario.isnot(None),
        PesquisaSatisfacao.comentario != ''
    ).distinct().order_by(PesquisaSatisfacao.especialidade).all()
    especialidades_disponiveis = [e[0] for e in especialidades_disponiveis]

    return render_template('admin_comentarios.html',
        comentarios=comentarios,
        total_registros=total_registros,
        pagina=pagina,
        total_paginas=total_paginas,
        por_pagina=por_pagina,
        especialidades_disponiveis=especialidades_disponiveis,
        # Filtros aplicados (para manter no form)
        especialidade_filtro=especialidade_filtro,
        nota_min=nota_min,
        nota_max=nota_max,
        usuario_filtro=usuario_filtro,
        data_inicio=data_inicio,
        data_fim=data_fim,
        busca=busca
    )
